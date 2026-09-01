import argparse
import json
import re
from pathlib import Path

from openpyxl import load_workbook


def normalize_data_type(value):
    mapping = {"INT": "INT16", "UINT": "UINT16"}
    if value not in mapping:
        raise ValueError(f"Unsupported data type: {value}")
    return mapping[value]


def export_model(input_path, output_path):
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    sheet = workbook["Sheet1"]
    points = []

    for row in sheet.iter_rows(min_row=3, values_only=True):
        if row[1] is None:
            continue
        address = int(row[1])
        points.append(
            {
                "index": int(row[0]),
                "address": address,
                "protocolOffset": address - 1,
                "source": str(row[2] or "").strip(),
                "access": str(row[3] or "").strip().lower(),
                "dataType": normalize_data_type(row[4]),
                "widthBits": int(str(row[5]).replace("位", "")),
                "name": str(row[6] or "").strip(),
                "scale": float(row[8]) if isinstance(row[8], (int, float)) and row[8] > 0 else 1,
            }
        )

    version_text = str(sheet["A1"].value or "")
    version_match = re.search(r"版本[：:]\s*([0-9]+(?:\.[0-9]+)*)", version_text)
    source_version = f"V{version_match.group(1)}" if version_match else version_text
    model = {
        "schemaVersion": 1,
        "sourceWorkbook": Path(input_path).name,
        "sourceSheet": "Sheet1",
        "sourceVersion": source_version,
        "addressConvention": "Workbook addresses are 1-based; protocolOffset is 0-based.",
        "points": points,
    }
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    Path(output_path).write_text(
        json.dumps(model, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("input")
    parser.add_argument("output")
    args = parser.parse_args()
    export_model(args.input, args.output)


if __name__ == "__main__":
    main()
